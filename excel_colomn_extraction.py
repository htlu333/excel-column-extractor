# ============== 标准库导入 ==============
import logging
import os
import sys
from collections import defaultdict
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict, Callable, Any
from enum import Enum, auto

# ============== 第三方库导入 ==============
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Fill, PatternFill
from openpyxl.utils import get_column_letter

# ============== 日志配置 ==============
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============== 异步任务框架 ==============
class TaskStatus(Enum):
    """任务状态枚举"""
    PENDING = auto()
    RUNNING = auto()
    COMPLETED = auto()
    FAILED = auto()
    CANCELLED = auto()


@dataclass
class TaskProgress:
    """任务进度信息"""
    current: int = 0
    total: int = 100
    message: str = ""
    status: TaskStatus = TaskStatus.PENDING
    
    @property
    def percentage(self) -> float:
        """计算完成百分比"""
        if self.total <= 0:
            return 0
        return min(100, int((self.current / self.total) * 100))


class CancellationToken:
    """取消令牌"""
    def __init__(self):
        self._cancelled = False
    
    def cancel(self):
        """请求取消"""
        self._cancelled = True
    
    def is_cancelled(self) -> bool:
        """检查是否已取消"""
        return self._cancelled
    
    def raise_if_cancelled(self):
        """如果已取消则抛出异常"""
        if self._cancelled:
            raise TaskCancelledException("任务已取消")


class TaskCancelledException(Exception):
    """任务取消异常"""
    pass


class AsyncTaskExecutor:
    """异步任务执行器"""
    def __init__(self, root: tk.Tk):
        self.root = root
        self.current_task = None
        self.cancellation_token = None
    
    def execute(
        self,
        task_func: Callable[[Callable[[int, int, str], None], CancellationToken], Any],
        on_complete: Optional[Callable[[Any], None]] = None,
        on_error: Optional[Callable[[Exception], None]] = None,
        on_progress: Optional[Callable[[TaskProgress], None]] = None,
        on_cancelled: Optional[Callable[[], None]] = None
    ):
        """
        执行异步任务
        
        Args:
            task_func: 任务函数，接收 (progress_callback, cancellation_token)
            on_complete: 完成回调
            on_error: 错误回调
            on_progress: 进度回调
            on_cancelled: 取消回调
        """
        self.cancellation_token = CancellationToken()
        
        def progress_callback(current: int, total: int, message: str):
            """进度回调包装"""
            progress = TaskProgress(current, total, message, TaskStatus.RUNNING)
            if on_progress:
                self.root.after(0, lambda: on_progress(progress))
        
        def run_task():
            try:
                result = task_func(progress_callback, self.cancellation_token)
                if not self.cancellation_token.is_cancelled():
                    if on_complete:
                        self.root.after(0, lambda: on_complete(result))
            except TaskCancelledException:
                if on_cancelled:
                    self.root.after(0, on_cancelled)
            except Exception as e:
                if on_error:
                    self.root.after(0, lambda: on_error(e))
        
        import threading
        thread = threading.Thread(target=run_task, daemon=True)
        thread.start()
        self.current_task = thread
    
    def cancel(self):
        """取消当前任务"""
        if self.cancellation_token:
            self.cancellation_token.cancel()


# ============== 数据模型 ==============
@dataclass
class ColumnInfo:
    """列信息"""
    name: str
    index: int  # 列索引（从1开始）
    letter: str  # 列字母（A, B, C...）
    file_index: int = 0  # 文件索引
    file_name: str = ""  # 文件名


@dataclass
class ExcelFileInfo:
    """Excel文件信息"""
    file_path: str
    sheet_name: str
    columns: List[ColumnInfo]
    total_rows: int
    file_index: int = 0  # 文件索引


# ============== 颜色配置 ==============
# 固定颜色列表（浅色背景）
FILE_COLORS = [
    "#E3F2FD",  # 浅蓝色
    "#E8F5E9",  # 浅绿色
    "#FFF3E0",  # 浅橙色
    "#FCE4EC",  # 浅粉色
    "#F3E5F5",  # 浅紫色
    "#E0F2F1",  # 浅青色
    "#FFF9C4",  # 浅黄色
    "#EFEBE9",  # 浅棕色
]


# ============== Excel分析器 ==============
class ExcelAnalyzer:
    """Excel文件分析器"""
    
    @staticmethod
    def load_file_info(file_path: str, sheet_name: Optional[str] = None, file_index: int = 0) -> ExcelFileInfo:
        """
        加载Excel文件信息
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，如果为None则使用第一个工作表
            file_index: 文件索引
        
        Returns:
            ExcelFileInfo对象
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
            
            # 选择工作表
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            
            file_name = os.path.basename(file_path)
            
            # 读取第一行作为列名
            columns: List[ColumnInfo] = []
            if sheet.max_row > 0:
                for col_idx, cell in enumerate(sheet[1], start=1):
                    column_name = str(cell.value) if cell.value is not None else f"列{col_idx}"
                    columns.append(ColumnInfo(
                        name=column_name,
                        index=col_idx,
                        letter=get_column_letter(col_idx),
                        file_index=file_index,
                        file_name=file_name
                    ))
            
            total_rows = sheet.max_row
            
            workbook.close()
            
            logger.info(f"成功加载文件: {file_path}, 工作表: {sheet.title}, 列数: {len(columns)}, 行数: {total_rows}")
            
            return ExcelFileInfo(
                file_path=file_path,
                sheet_name=sheet.title,
                columns=columns,
                total_rows=total_rows,
                file_index=file_index
            )
        except Exception as e:
            logger.error(f"加载Excel文件失败: {e}")
            raise
    
    @staticmethod
    def extract_columns(
        file_path: str,
        sheet_name: str,
        selected_columns: List[str],
        output_path: str,
        progress_callback: Callable[[int, int, str], None],
        cancellation_token: CancellationToken
    ) -> str:
        """
        提取选定的列并保存为新文件（保留格式）
        
        Args:
            file_path: 源文件路径
            sheet_name: 工作表名称
            selected_columns: 选定的列名列表
            output_path: 输出文件路径
            progress_callback: 进度回调
            cancellation_token: 取消令牌
        
        Returns:
            输出文件路径
        """
        try:
            # 打开源文件
            source_workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
            source_sheet = source_workbook[sheet_name]
            
            # 创建新工作簿
            output_workbook = openpyxl.Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "合并列"
            
            # 获取选定的列索引
            column_map: Dict[str, int] = {}
            for col_idx, cell in enumerate(source_sheet[1], start=1):
                col_name = str(cell.value) if cell.value is not None else f"列{col_idx}"
                if col_name in selected_columns:
                    column_map[col_name] = col_idx
            
            # 确定输出列的顺序
            output_col_idx = 1
            for col_name in selected_columns:
                if col_name in column_map:
                    source_col_idx = column_map[col_name]
                    source_col_letter = get_column_letter(source_col_idx)
                    
                    # 复制数据
                    total_rows = source_sheet.max_row
                    for row_idx in range(1, total_rows + 1):
                        cancellation_token.raise_if_cancelled()
                        
                        source_cell = source_sheet[f"{source_col_letter}{row_idx}"]
                        output_cell = output_sheet.cell(row=row_idx, column=output_col_idx)
                        
                        # 复制值
                        output_cell.value = source_cell.value
                        
                        # 复制格式
                        if source_cell.has_style:
                            output_cell.font = Font(
                                name=source_cell.font.name if source_cell.font else None,
                                size=source_cell.font.size if source_cell.font else None,
                                bold=source_cell.font.bold if source_cell.font else False,
                                italic=source_cell.font.italic if source_cell.font else False,
                                color=source_cell.font.color if source_cell.font else None
                            )
                            output_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal if source_cell.alignment else None,
                                vertical=source_cell.alignment.vertical if source_cell.alignment else None,
                                wrap_text=source_cell.alignment.wrap_text if source_cell.alignment else False
                            )
                            if source_cell.border:
                                output_cell.border = Border(
                                    left=source_cell.border.left,
                                    right=source_cell.border.right,
                                    top=source_cell.border.top,
                                    bottom=source_cell.border.bottom
                                )
                            if source_cell.fill:
                                output_cell.fill = PatternFill(
                                    fill_type=source_cell.fill.fill_type,
                                    start_color=source_cell.fill.start_color,
                                    end_color=source_cell.fill.end_color
                                )
                            output_cell.number_format = source_cell.number_format
                        
                        # 更新进度
                        if row_idx % 100 == 0:
                            progress = int((row_idx / total_rows) * 100)
                            progress_callback(
                                row_idx * output_col_idx,
                                total_rows * len(selected_columns),
                                f"正在复制列 '{col_name}' ({row_idx}/{total_rows})"
                            )
                    
                    # 调整列宽
                    if source_sheet.column_dimensions[source_col_letter].width:
                        output_sheet.column_dimensions[get_column_letter(output_col_idx)].width = \
                            source_sheet.column_dimensions[source_col_letter].width
                    
                    output_col_idx += 1
            
            # 保存文件
            progress_callback(100, 100, "正在保存文件...")
            output_workbook.save(output_path)
            output_workbook.close()
            source_workbook.close()
            
            logger.info(f"成功提取列到: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"提取列失败: {e}")
            raise
    
    @staticmethod
    def merge_multiple_files(
        file_infos: List[ExcelFileInfo],
        selected_cols_info: List[Tuple[int, str]],
        reference_column_map: Dict[str, int],
        output_path: str,
        progress_callback: Callable[[int, int, str], None],
        cancellation_token: CancellationToken
    ) -> str:
        """
        合并多个文件的列
        
        Args:
            file_infos: 文件信息列表
            selected_cols_info: 选定的列信息列表，格式为 [(file_index, column_name), ...]
            reference_column_map: 参照列映射，格式为 {column_name: reference_file_index}
            output_path: 输出文件路径
            progress_callback: 进度回调
            cancellation_token: 取消令牌
        
        Returns:
            输出文件路径
        """
        try:
            # 打开所有源文件
            source_workbooks: List[Any] = []
            source_sheets: List[Any] = []
            
            for file_info in file_infos:
                wb = openpyxl.load_workbook(file_info.file_path, read_only=False, data_only=False)
                source_workbooks.append(wb)
                source_sheets.append(wb[file_info.sheet_name])
            
            # 创建新工作簿
            output_workbook = openpyxl.Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "合并列"
            
            # 建立列映射：{file_index: {column_name: column_index}}
            column_maps: Dict[int, Dict[str, int]] = {}
            for file_index, file_info in enumerate(file_infos):
                column_map: Dict[str, int] = {}
                sheet = source_sheets[file_index]
                for col_idx, cell in enumerate(sheet[1], start=1):
                    col_name = str(cell.value) if cell.value is not None else f"列{col_idx}"
                    column_map[col_name] = col_idx
                column_maps[file_index] = column_map
            
            # 读取所有文件的数据
            file_data: Dict[int, Dict[int, List[Any]]] = {}  # {file_index: {row: [cell_values]}}
            max_rows: Dict[int, int] = {}
            
            for file_index, sheet in enumerate(source_sheets):
                max_row = sheet.max_row
                max_rows[file_index] = max_row
                file_data[file_index] = {}
                
                for row_idx in range(1, max_row + 1):
                    row_data = []
                    for col_idx, cell in enumerate(sheet[row_idx], start=1):
                        row_data.append(cell)
                    file_data[file_index][row_idx] = row_data
            
            # 处理参照列，建立主键映射
            reference_keys: Dict[str, List[Any]] = {}  # {column_name: [key_values]}
            reference_key_to_row: Dict[str, Dict[Any, int]] = {}  # {column_name: {key_value: row_index}}
            
            for column_name, ref_file_index in reference_column_map.items():
                if column_name not in column_maps[ref_file_index]:
                    continue
                
                ref_col_idx = column_maps[ref_file_index][column_name]
                ref_sheet = source_sheets[ref_file_index]
                keys = []
                key_to_row = {}
                
                for row_idx in range(2, max_rows[ref_file_index] + 1):  # 从第2行开始（跳过标题行）
                    cell = ref_sheet.cell(row=row_idx, column=ref_col_idx)
                    key_value = cell.value
                    if key_value not in key_to_row:
                        keys.append(key_value)
                        key_to_row[key_value] = row_idx
                
                reference_keys[column_name] = keys
                reference_key_to_row[column_name] = key_to_row
            
            # 按文件顺序组织列
            columns_by_file: Dict[int, List[Tuple[int, str]]] = defaultdict(list)
            for file_index, column_name in selected_cols_info:
                columns_by_file[file_index].append((file_index, column_name))
            
            # 确定最终的行数和行映射
            # 对于有参照列的，使用参照列的行数；否则使用各自文件的行数
            final_row_count = 1  # 标题行
            row_mapping: Dict[int, Dict[int, int]] = {}  # {file_index: {source_row: target_row}}
            
            for file_index in columns_by_file.keys():
                file_row_mapping: Dict[int, int] = {}
                
                # 检查该文件的列是否有参照列
                has_ref_column = False
                ref_column_name = None
                
                for _, column_name in columns_by_file[file_index]:
                    if column_name in reference_column_map:
                        has_ref_column = True
                        ref_column_name = column_name
                        break
                
                if has_ref_column and ref_column_name:
                    # 使用参照列的行映射
                    ref_file_index = reference_column_map[ref_column_name]
                    ref_keys = reference_keys[ref_column_name]
                    
                    # 建立该文件到参照列的映射
                    file_col_idx = column_maps[file_index].get(ref_column_name)
                    if file_col_idx:
                        file_sheet = source_sheets[file_index]
                        file_key_to_row: Dict[Any, int] = {}
                        
                        for row_idx in range(2, max_rows[file_index] + 1):
                            cell = file_sheet.cell(row=row_idx, column=file_col_idx)
                            key_value = cell.value
                            if key_value not in file_key_to_row:
                                file_key_to_row[key_value] = row_idx
                        
                        # 按参照列的顺序映射
                        target_row = 2
                        for ref_key in ref_keys:
                            if ref_key in file_key_to_row:
                                file_row_mapping[file_key_to_row[ref_key]] = target_row
                            else:
                                # 该文件没有这个key，跳过（会在后面追加空行）
                                pass
                            target_row += 1
                        
                        # 追加该文件独有的key
                        for file_key, file_row in file_key_to_row.items():
                            if file_key not in reference_key_to_row[ref_column_name]:
                                file_row_mapping[file_row] = target_row
                                target_row += 1
                        
                        final_row_count = max(final_row_count, target_row)
                    else:
                        # 该文件没有参照列，直接映射
                        target_row = 2
                        for row_idx in range(2, max_rows[file_index] + 1):
                            file_row_mapping[row_idx] = target_row
                            target_row += 1
                        final_row_count = max(final_row_count, target_row)
                else:
                    # 没有参照列，直接映射
                    target_row = 2
                    for row_idx in range(2, max_rows[file_index] + 1):
                        file_row_mapping[row_idx] = target_row
                        target_row += 1
                    final_row_count = max(final_row_count, target_row)
                
                row_mapping[file_index] = file_row_mapping
            
            # 合并数据
            output_col_idx = 1
            total_cols = len(selected_cols_info)
            processed_cols = 0
            
            # 按文件顺序输出列
            for file_index in sorted(columns_by_file.keys()):
                for _, column_name in columns_by_file[file_index]:
                    cancellation_token.raise_if_cancelled()
                    processed_cols += 1
                    
                    source_col_idx = column_maps[file_index].get(column_name)
                    if not source_col_idx:
                        continue
                    
                    source_sheet = source_sheets[file_index]
                    source_col_letter = get_column_letter(source_col_idx)
                    file_row_map = row_mapping[file_index]
                    
                    # 写入标题
                    header_cell = source_sheet.cell(row=1, column=source_col_idx)
                    output_header = output_sheet.cell(row=1, column=output_col_idx)
                    output_header.value = header_cell.value
                    if header_cell.has_style:
                        ExcelAnalyzer._copy_cell_style(header_cell, output_header)
                    
                    # 写入数据
                    for source_row, target_row in file_row_map.items():
                        cancellation_token.raise_if_cancelled()
                        
                        source_cell = source_sheet.cell(row=source_row, column=source_col_idx)
                        output_cell = output_sheet.cell(row=target_row, column=output_col_idx)
                        
                        output_cell.value = source_cell.value
                        if source_cell.has_style:
                            ExcelAnalyzer._copy_cell_style(source_cell, output_cell)
                    
                    # 调整列宽
                    if source_sheet.column_dimensions[source_col_letter].width:
                        output_sheet.column_dimensions[get_column_letter(output_col_idx)].width = \
                            source_sheet.column_dimensions[source_col_letter].width
                    
                    # 更新进度
                    progress_callback(
                        processed_cols,
                        total_cols,
                        f"正在合并列 '{column_name}' ({processed_cols}/{total_cols})"
                    )
                    
                    output_col_idx += 1
            
            # 关闭所有源文件
            for wb in source_workbooks:
                wb.close()
            
            # 保存文件
            progress_callback(100, 100, "正在保存文件...")
            output_workbook.save(output_path)
            output_workbook.close()
            
            logger.info(f"成功合并文件到: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"合并文件失败: {e}")
            raise
    
    @staticmethod
    def _copy_cell_style(source_cell: Any, target_cell: Any):
        """复制单元格样式"""
        if source_cell.has_style:
            target_cell.font = Font(
                name=source_cell.font.name if source_cell.font else None,
                size=source_cell.font.size if source_cell.font else None,
                bold=source_cell.font.bold if source_cell.font else False,
                italic=source_cell.font.italic if source_cell.font else False,
                color=source_cell.font.color if source_cell.font else None
            )
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal if source_cell.alignment else None,
                vertical=source_cell.alignment.vertical if source_cell.alignment else None,
                wrap_text=source_cell.alignment.wrap_text if source_cell.alignment else False
            )
            if source_cell.border:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            target_cell.number_format = source_cell.number_format


# ============== GUI界面 ==============
class StyledLabelFrame(ttk.LabelFrame):
    """带样式的LabelFrame"""
    def __init__(self, parent, **kwargs):
        kwargs.setdefault("style", "Styled.TLabelframe")
        super().__init__(parent, **kwargs)


class ModernButton(ttk.Button):
    """现代化按钮"""
    def __init__(self, parent, **kwargs):
        kwargs.setdefault("style", "Modern.TButton")
        super().__init__(parent, **kwargs)


class ReferenceColumnDialog:
    """参照列选择对话框"""
    def __init__(self, parent, column_name: str, file_options: List[Tuple[int, str]]):
        """
        初始化参照列选择对话框
        
        Args:
            parent: 父窗口
            column_name: 列名
            file_options: 文件选项列表，格式为 [(file_index, file_name), ...]
        """
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("选择参照列")
        self.dialog.geometry("500x300")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self.selected_file_index: Optional[int] = None
        self._setup_ui(column_name, file_options)
        self._center_window()
    
    def _center_window(self):
        """居中显示窗口"""
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry(f"{width}x{height}+{x}+{y}")
    
    def _setup_ui(self, column_name: str, file_options: List[Tuple[int, str]]):
        """设置UI"""
        frame = ttk.Frame(self.dialog, padding="20")
        frame.pack(fill="both", expand=True)
        
        # 提示信息
        info_label = tk.Label(
            frame,
            text=f"检测到多个文件包含列 '{column_name}'，\n请选择其中一个文件的列作为参照列（主键）：",
            font=("Microsoft YaHei", 10),
            justify="left"
        )
        info_label.pack(pady=(0, 15))
        
        # 文件选项列表
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        self.file_var = tk.IntVar(value=file_options[0][0] if file_options else -1)
        
        for file_index, file_name in file_options:
            file_color = FILE_COLORS[file_index % len(FILE_COLORS)]
            option_frame = tk.Frame(list_frame, bg=file_color)
            option_frame.pack(fill="x", padx=10, pady=2)
            
            rb = tk.Radiobutton(
                option_frame,
                text=f"[文件{file_index + 1}] {file_name}",
                variable=self.file_var,
                value=file_index,
                font=("Microsoft YaHei", 10),
                bg=file_color
            )
            rb.pack(anchor="w", padx=10, pady=5)
        
        # 按钮
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x")
        
        ModernButton(btn_frame, text="确定", command=self._on_ok).pack(side="right", padx=(10, 0))
        ModernButton(btn_frame, text="取消", command=self._on_cancel).pack(side="right")
    
    def _on_ok(self):
        """确定按钮"""
        self.selected_file_index = self.file_var.get()
        self.dialog.destroy()
    
    def _on_cancel(self):
        """取消按钮"""
        self.selected_file_index = None
        self.dialog.destroy()
    
    def get_result(self) -> Optional[int]:
        """获取选择结果"""
        self.dialog.wait_window()
        return self.selected_file_index


class ProgressDialog:
    """进度对话框"""
    def __init__(self, parent, title: str = "处理中..."):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("400x150")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.protocol("WM_DELETE_WINDOW", self._on_cancel)
        
        self.cancelled = False
        self._center_window()
        self._setup_ui()
    
    def _center_window(self):
        """居中显示窗口"""
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry(f"{width}x{height}+{x}+{y}")
    
    def _setup_ui(self):
        """设置UI"""
        frame = ttk.Frame(self.dialog, padding="20")
        frame.pack(fill="both", expand=True)
        
        self.status_label = tk.Label(frame, text="准备中...", font=("Microsoft YaHei", 10))
        self.status_label.pack(pady=(0, 10))
        
        self.progress_bar = ttk.Progressbar(frame, mode="determinate", length=300)
        self.progress_bar.pack(pady=(0, 10))
        
        self.percentage_label = tk.Label(frame, text="0%", font=("Microsoft YaHei", 9))
        self.percentage_label.pack(pady=(0, 10))
        
        self.cancel_btn = ModernButton(frame, text="取消", command=self._on_cancel)
        self.cancel_btn.pack()
    
    def _on_cancel(self):
        """取消操作"""
        self.cancelled = True
        self.dialog.destroy()
    
    def update_progress(self, progress: TaskProgress):
        """更新进度"""
        if self.dialog.winfo_exists():
            self.status_label.config(text=progress.message)
            self.progress_bar["value"] = progress.percentage
            self.percentage_label.config(text=f"{progress.percentage}%")
            self.dialog.update()


class MainUI:
    """主界面"""
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Excel列提取工具")
        self.root.geometry("480x800")
        
        # DPI适配
        try:
            
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass
        
        self.file_infos: List[ExcelFileInfo] = []  # 多文件信息列表
        self.column_vars: Dict[str, tk.BooleanVar] = {}  # 列选择变量，key为"file_index:column_name"
        self.async_executor = AsyncTaskExecutor(root)
        self.last_output_folder: Optional[str] = None  # 保存最后输出的文件夹路径
        self.last_output_file: Optional[str] = None  # 保存最后输出的文件路径
        
        self._setup_theme()
        self._setup_ui()
    
    def _setup_theme(self):
        """设置主题"""
        s = ttk.Style()
        try:
            s.theme_use("xpnative")
        except Exception:
            pass
        
        s.configure(
            "Modern.TButton",
            font=("Microsoft YaHei", 10, "bold"),
            padding=(15, 6)
        )
        
        s.configure(
            "Styled.TLabelframe",
            font=("Microsoft YaHei", 10, "bold")
        )
        
        s.configure(
            "Column.TCheckbutton",
            font=("Microsoft YaHei", 10)
        )
    
    def _setup_ui(self):
        """设置UI"""
        # 主容器
        main = ttk.Frame(self.root, padding="12")
        main.pack(fill="both", expand=True)
        
        # 标题
        title_label = tk.Label(
            main,
            text="Excel列提取工具",
            font=("Microsoft YaHei", 18, "bold"),
            fg="#4CAAB9"
        )
        title_label.pack(pady=(0, 20))
        
        # 文件选择区域
        file_frame = StyledLabelFrame(main, text="📂 文件选择", padding="10")
        file_frame.pack(fill="x", pady=(0, 15))
        
        file_inner = ttk.Frame(file_frame)
        file_inner.pack(fill="x", pady=(0, 10))
        
        ModernButton(file_inner, text="选择Excel文件（可多选）", command=self._on_select_files).pack(side="left", padx=(0, 10))
        
        # 自动打开文件选项
        self.auto_open_file_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            file_inner,
            text="自动打开结果文件",
            variable=self.auto_open_file_var,
            style="Column.TCheckbutton"
        ).pack(side="left")
        
        # 文件列表显示区域
        list_frame = ttk.Frame(file_frame)
        list_frame.pack(fill="both", expand=True)
        
        # 文件列表（带滚动条）
        file_list_container = ttk.Frame(list_frame)
        file_list_container.pack(fill="both", expand=True)
        
        file_scrollbar = ttk.Scrollbar(file_list_container)
        file_scrollbar.pack(side="right", fill="y")
        
        self.file_listbox = tk.Listbox(
            file_list_container,
            yscrollcommand=file_scrollbar.set,
            height=4,
            font=("Microsoft YaHei", 9)
        )
        self.file_listbox.pack(side="left", fill="both", expand=True)
        file_scrollbar.config(command=self.file_listbox.yview)
        
        # 移除文件按钮
        remove_btn_frame = ttk.Frame(file_frame)
        remove_btn_frame.pack(fill="x", pady=(5, 0))
        ModernButton(remove_btn_frame, text="移除选中文件", command=self._on_remove_file).pack(side="left")
        
        # 列选择区域
        column_frame = StyledLabelFrame(main, text="⚙ 列选择", padding="10")
        column_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # 全选/全不选按钮（放在列选择框上面）
        select_buttons_frame = ttk.Frame(column_frame)
        select_buttons_frame.pack(fill="x", pady=(0, 10))
        
        # 创建小按钮
        small_btn_style = ttk.Style()
        small_btn_style.configure(
            "Small.TButton",
            font=("Microsoft YaHei", 9),
            padding=(8, 3)
        )
        
        ttk.Button(select_buttons_frame, text="全选", command=self._on_select_all, style="Small.TButton").pack(side="left", padx=(0, 5))
        ttk.Button(select_buttons_frame, text="全不选", command=self._on_deselect_all, style="Small.TButton").pack(side="left")
        
        # 列列表容器（带滚动条）
        list_container = ttk.Frame(column_frame)
        list_container.pack(fill="both", expand=True)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(list_container)
        scrollbar.pack(side="right", fill="y")
        
        # 列列表（Canvas + Frame实现滚动）
        self.column_canvas = tk.Canvas(
            list_container,
            yscrollcommand=scrollbar.set,
            bg="white",
            highlightthickness=0
        )
        self.column_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.column_canvas.yview)
        
        # 内部Frame用于放置复选框
        self.column_inner_frame = ttk.Frame(self.column_canvas)
        self.column_canvas_window = self.column_canvas.create_window(
            (0, 0),
            window=self.column_inner_frame,
            anchor="nw"
        )
        
        # 绑定滚动区域更新
        self.column_inner_frame.bind(
            "<Configure>",
            lambda e: self.column_canvas.configure(scrollregion=self.column_canvas.bbox("all"))
        )
        self.column_canvas.bind(
            "<Configure>",
            lambda e: self._on_canvas_configure()
        )
        
        # 绑定鼠标滚轮事件
        self._bind_mousewheel()
        
        # 操作按钮区域（底部）
        action_frame = ttk.Frame(main)
        action_frame.pack(fill="x", pady=(0, 15))
        
        self.export_btn = ModernButton(action_frame, text="📄 输出拆分合并Excel", command=self._on_export)
        self.export_btn.pack(side="left", padx=(0, 10))
        
        self.open_folder_btn = ModernButton(action_frame, text="📂 打开结果文件夹", command=self._on_open_folder)
        self.open_folder_btn.pack(side="left")
        
        # 状态栏
        status = tk.Frame(self.root, bg="#E8E8E8", height=25)
        status.pack(fill="x", side="bottom")
        self.status_var = tk.StringVar(value="🟢 就绪")
        tk.Label(
            status,
            textvariable=self.status_var,
            bg="#E8E8E8",
            fg="#555",
            font=("Microsoft YaHei", 10)
        ).pack(anchor="w", padx=10)
    
    def _on_canvas_configure(self):
        """Canvas大小改变时调整内部Frame宽度"""
        canvas_width = self.column_canvas.winfo_width()
        self.column_canvas.itemconfig(self.column_canvas_window, width=canvas_width)
    
    def _bind_mousewheel(self):
        """绑定鼠标滚轮事件"""
        def _on_mousewheel(event):
            """鼠标滚轮事件处理"""
            # 检查鼠标是否在canvas区域内
            try:
                x = self.column_canvas.winfo_pointerx() - self.column_canvas.winfo_rootx()
                y = self.column_canvas.winfo_pointery() - self.column_canvas.winfo_rooty()
                if 0 <= x < self.column_canvas.winfo_width() and 0 <= y < self.column_canvas.winfo_height():
                    if sys.platform.startswith("win"):
                        # Windows平台
                        self.column_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                    else:
                        # Linux/Mac平台
                        if event.num == 4:
                            self.column_canvas.yview_scroll(-1, "units")
                        elif event.num == 5:
                            self.column_canvas.yview_scroll(1, "units")
            except:
                pass
        
        # 使用bind_all全局绑定，但只在canvas区域内响应
        # 这样不会因为鼠标离开而失效
        if sys.platform.startswith("win"):
            self.root.bind_all("<MouseWheel>", _on_mousewheel)
        else:
            self.root.bind_all("<Button-4>", _on_mousewheel)
            self.root.bind_all("<Button-5>", _on_mousewheel)
    
    def _on_select_files(self):
        """选择多个文件"""
        file_paths = filedialog.askopenfilenames(
            title="选择Excel文件（可多选）",
            filetypes=[("Excel文件", "*.xlsx *.xlsm *.xls")]
        )
        
        if not file_paths:
            return
        
        try:
            self.status_var.set("⏳ 正在加载文件...")
            self.root.update()
            
            new_files_count = 0
            for file_path in file_paths:
                # 检查文件是否已存在
                if any(info.file_path == file_path for info in self.file_infos):
                    continue
                
                # 加载文件信息
                file_index = len(self.file_infos)
                file_info = ExcelAnalyzer.load_file_info(file_path, file_index=file_index)
                self.file_infos.append(file_info)
                
                # 添加到文件列表
                self.file_listbox.insert(tk.END, f"[文件{file_index + 1}] {os.path.basename(file_path)}")
                new_files_count += 1
            
            # 更新列列表
            self._update_column_list()
            
            total_columns = sum(len(info.columns) for info in self.file_infos)
            self.status_var.set(f"✅ 已加载 {new_files_count} 个文件，共 {len(self.file_infos)} 个文件，{total_columns} 列")
            
        except Exception as e:
            logger.error(f"加载文件失败: {e}")
            messagebox.showerror("错误", f"加载文件失败：{e}")
            self.status_var.set("❌ 加载失败")
    
    def _on_remove_file(self):
        """移除选中的文件"""
        selected_indices = self.file_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("提示", "请先选择要移除的文件")
            return
        
        # 从后往前删除，避免索引变化
        for index in reversed(selected_indices):
            if 0 <= index < len(self.file_infos):
                self.file_infos.pop(index)
                self.file_listbox.delete(index)
        
        # 重新分配文件索引
        for i, file_info in enumerate(self.file_infos):
            file_info.file_index = i
            for col in file_info.columns:
                col.file_index = i
        
        # 更新文件列表显示
        self.file_listbox.delete(0, tk.END)
        for i, file_info in enumerate(self.file_infos):
            self.file_listbox.insert(tk.END, f"[文件{i + 1}] {os.path.basename(file_info.file_path)}")
        
        # 更新列列表
        self._update_column_list()
        
        total_columns = sum(len(info.columns) for info in self.file_infos)
        self.status_var.set(f"✅ 已移除，剩余 {len(self.file_infos)} 个文件，{total_columns} 列")
    
    def _update_column_list(self):
        """更新列列表"""
        # 清空现有复选框
        for widget in self.column_inner_frame.winfo_children():
            widget.destroy()
        self.column_vars.clear()
        
        if not self.file_infos:
            return
        
        # 按文件顺序创建复选框
        for file_info in self.file_infos:
            file_color = FILE_COLORS[file_info.file_index % len(FILE_COLORS)]
            
            for col_info in file_info.columns:
                # 使用 "file_index:column_name" 作为唯一标识
                var_key = f"{col_info.file_index}:{col_info.name}"
                var = tk.BooleanVar(value=False)
                self.column_vars[var_key] = var
                
                # 创建带背景色的Frame
                col_frame = tk.Frame(self.column_inner_frame, bg=file_color)
                col_frame.pack(fill="x", padx=5, pady=1)
                
                checkbutton = ttk.Checkbutton(
                    col_frame,
                    text=f"[文件{col_info.file_index + 1}] {col_info.letter}: {col_info.name}",
                    variable=var,
                    style="Column.TCheckbutton"
                )
                checkbutton.pack(anchor="w", padx=5, pady=2)
        
        # 更新滚动区域
        self.column_inner_frame.update_idletasks()
        self.column_canvas.configure(scrollregion=self.column_canvas.bbox("all"))
    
    def _on_select_all(self):
        """全选"""
        for var in self.column_vars.values():
            var.set(True)
    
    def _on_deselect_all(self):
        """全不选"""
        for var in self.column_vars.values():
            var.set(False)
    
    def _on_export(self):
        """导出合并Excel"""
        if not self.file_infos:
            messagebox.showwarning("提示", "请先选择Excel文件")
            return
        
        # 获取选定的列（格式：file_index:column_name）
        selected_columns = [key for key, var in self.column_vars.items() if var.get()]
        
        if not selected_columns:
            messagebox.showwarning("提示", "请至少选择一列")
            return
        
        # 解析选定的列，转换为 (file_index, column_name) 格式
        selected_cols_info: List[Tuple[int, str]] = []
        for key in selected_columns:
            parts = key.split(":", 1)
            if len(parts) == 2:
                file_index = int(parts[0])
                column_name = parts[1]
                selected_cols_info.append((file_index, column_name))
        
        # 检测相同列
        column_name_to_files: Dict[str, List[Tuple[int, str]]] = defaultdict(list)
        for file_index, column_name in selected_cols_info:
            column_name_to_files[column_name].append((file_index, column_name))
        
        # 找出有相同列名的列
        duplicate_columns: Dict[str, List[Tuple[int, str]]] = {
            col_name: files for col_name, files in column_name_to_files.items() 
            if len(files) > 1
        }
        
        # 参照列映射：{列名: 参照文件索引}
        reference_column_map: Dict[str, int] = {}
        
        # 如果有相同列，弹出对话框让用户选择参照列
        if duplicate_columns:
            for column_name, file_list in duplicate_columns.items():
                file_options = [
                    (file_index, os.path.basename(self.file_infos[file_index].file_path)) 
                    for file_index, _ in file_list
                ]
                dialog = ReferenceColumnDialog(self.root, column_name, file_options)
                ref_file_index = dialog.get_result()
                
                if ref_file_index is None:
                    # 用户取消了
                    return
                
                reference_column_map[column_name] = ref_file_index
        
        # 生成默认文件名：第一个文件名 + "拆分合并"
        if self.file_infos:
            first_file = self.file_infos[0]
            source_dir = os.path.dirname(first_file.file_path)
            source_name = os.path.splitext(os.path.basename(first_file.file_path))[0]
            default_filename = f"{source_name}拆分合并.xlsx"
        else:
            default_filename = "拆分合并.xlsx"
            source_dir = os.getcwd()
        
        # 选择输出路径
        output_path = filedialog.asksaveasfilename(
            title="保存合并Excel",
            defaultextension=".xlsx",
            initialfile=default_filename,
            initialdir=source_dir,
            filetypes=[("Excel文件", "*.xlsx")]
        )
        
        if not output_path:
            return
        
        # 创建进度对话框
        progress_dialog = ProgressDialog(self.root, "正在合并列...")
        
        # 禁用按钮
        self._set_buttons_state(False)
        self.status_var.set("⏳ 处理中...")
        
        # 执行异步任务
        def task_func(progress_callback, cancellation_token):
            return ExcelAnalyzer.merge_multiple_files(
                self.file_infos,
                selected_cols_info,
                reference_column_map,
                output_path,
                progress_callback,
                cancellation_token
            )
        
        def on_complete(result):
            progress_dialog.dialog.destroy()
            self._set_buttons_state(True)
            self.status_var.set("✅ 完成")
            
            # 保存输出文件夹路径
            self.last_output_folder = os.path.dirname(result)
            self.last_output_file = result  # 保存最后输出的文件路径
            
            messagebox.showinfo("成功", f"已成功合并 {len(selected_columns)} 列到:\n{result}")
            
            # 根据复选框决定是否自动打开文件
            if self.auto_open_file_var.get():
                self._open_file(result)
        
        def on_error(error):
            progress_dialog.dialog.destroy()
            self._set_buttons_state(True)
            self.status_var.set("❌ 失败")
            messagebox.showerror("错误", f"合并失败：{error}")
        
        def on_progress(progress):
            if progress_dialog.dialog.winfo_exists():
                progress_dialog.update_progress(progress)
        
        def on_cancelled():
            progress_dialog.dialog.destroy()
            self._set_buttons_state(True)
            self.status_var.set("⚠️ 已取消")
        
        self.async_executor.execute(
            task_func=task_func,
            on_complete=on_complete,
            on_error=on_error,
            on_progress=on_progress,
            on_cancelled=on_cancelled
        )
    
    def _set_buttons_state(self, enabled: bool):
        """设置按钮状态"""
        state = "normal" if enabled else "disabled"
        if self.export_btn:
            self.export_btn.configure(state=state)
        if self.open_folder_btn:
            self.open_folder_btn.configure(state=state)
    
    def _open_folder(self, file_path: Optional[str] = None):
        """打开文件夹"""
        try:
            if file_path:
                folder_path = os.path.dirname(file_path)
            elif self.last_output_folder:
                folder_path = self.last_output_folder
            else:
                messagebox.showwarning("提示", "还没有生成过文件")
                return
            
            if sys.platform.startswith("win"):
                os.startfile(folder_path)
            elif sys.platform.startswith("darwin"):
                os.system(f'open "{folder_path}"')
            else:
                os.system(f'xdg-open "{folder_path}"')
        except Exception as e:
            logger.error(f"打开文件夹失败: {e}")
            messagebox.showerror("错误", f"无法打开文件夹：{e}")
    
    def _open_file(self, file_path: Optional[str] = None):
        """打开Excel文件"""
        try:
            path = file_path if file_path else self.last_output_file
            if not path:
                messagebox.showwarning("提示", "还没有生成过文件")
                return
            
            if sys.platform.startswith("win"):
                os.startfile(path)
            elif sys.platform.startswith("darwin"):
                os.system(f'open "{path}"')
            else:
                os.system(f'xdg-open "{path}"')
        except Exception as e:
            logger.error(f"打开文件失败: {e}")
            messagebox.showerror("错误", f"无法打开文件：{e}")
    
    def _on_open_folder(self):
        """打开结果文件夹按钮点击事件"""
        self._open_folder()


# ============== 主程序入口 ==============
def main():
    """主函数"""
    root = tk.Tk()
    app = MainUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

